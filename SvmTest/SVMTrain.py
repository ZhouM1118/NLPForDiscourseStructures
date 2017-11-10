import openpyxl
import config
import math
from sklearn import svm
from sklearn.metrics import accuracy_score
from sklearn.metrics import hamming_loss
from datetime import datetime

configs = config.configs

# 记录程序开始执行时间
start = datetime.now()

readBook = openpyxl.load_workbook(configs['extractFeaturesPath'])
readSheet = readBook.active

def getFeatureVector():
    featureVectorList = []
    sentenceTag = []
    for i in range(readSheet.max_row - 1):

        featureVector = []
        sentenceTag.append(readSheet.cell(row = i + 2, column = 5).value)
        # 先添加段落特征
        featureVector.append(float(readSheet.cell(row = i + 2, column = 1).value.strip().split('-')[1]))
        for j in range(readSheet.max_column - 5):
            if isinstance(readSheet.cell(row = i + 2, column = j + 6).value, str):
                featureVector.append(float(readSheet.cell(row=i + 2, column=j + 6).value))
            else:
                featureVector.append(readSheet.cell(row = i + 2, column = j + 6).value)

        featureVectorList.append(normalizeFeatureBySigmoid(featureVector))
    return featureVectorList, sentenceTag

# 归一化处理，使用sigmoid函数
# 2017.11.10 准确率 0.205202312139
# ('6-5', 195), ('1-5', 108), ('4-5', 95), ('10-5', 74), ('12-5', 68), ('7-5', 66),
# ('13-5', 60), ('9-5', 51), ('2-5', 45), ('11-5', 44), ('3-5', 17), ('8-5', 2)
def normalizeFeatureBySigmoid(featureVector):
    normalizeFeatureVector = []
    for feature in featureVector:
        normalizeFeatureVector.append(1.0 / (1 + math.exp(-float(feature))))
    return normalizeFeatureVector

# 归一化处理，使用(0,1)标准化
# 2017.11.10 准确率 0.401734104046
# ('1-5', 90), ('4-5', 88), ('10-5', 67), ('12-5', 65), ('7-5', 52), ('9-5', 49),
#  ('11-5', 41), ('2-5', 40), ('13-6', 34), ('13-5', 25), ('7-6', 11), ('3-5', 10),
#  ('4-6', 6), ('10-6', 6), ('1-6', 6), ('3-6', 5), ('5-6', 3), ('12-6', 3), ('7-1', 3),
#  ('2-6', 3), ('9-6', 2), ('3-1', 2), ('8-5', 2), ('11-6', 2), ('2-1', 2), ('13-1', 1),
#  ('4-1', 1), ('10-1', 1), ('11-1', 1)
def normalizeFeatureByMaxMin(featureVector):
    normalizeFeatureVector = []
    maxNum = max(featureVector)
    minNum = min(featureVector)
    for feature in featureVector:
        normalizeFeatureVector.append((feature - minNum) / (maxNum - minNum))
    return normalizeFeatureVector

# 训练并测试训练集的准确度
def doTrain(X, Y):
    clf = svm.SVC()
    clf.fit(X, Y)

    Y_pred = clf.predict(X)
    print(accuracy_score(Y, Y_pred))
    print(Y)
    print(Y_pred)
    print('Y len is ', len(Y),'Y_pred len is ', len(Y_pred))

    compare = {}
    j = 0
    for i in range(len(Y)):
        if Y[i] == 5:
            j += 1
        YAndYpred = str(Y[i]) + '-' + str(Y_pred[i])
        # print(YAndYpred)
        if(Y[i] != Y_pred[i]):
            if YAndYpred not in compare:
                compare[YAndYpred] = 1
            else:
                compare[YAndYpred] = compare[YAndYpred] + 1

    print(sorted(compare.items(), key=lambda d: -d[1]))

params = getFeatureVector()
doTrain(params[0], params[1])

# 计算程序运行总时间(秒)
elapsed = (datetime.now() - start).seconds
print('Time used : ', elapsed)