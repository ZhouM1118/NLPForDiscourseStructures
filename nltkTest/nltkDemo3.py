import openpyxl
import config
configs = config.configs

writeBook = openpyxl.load_workbook(configs['extractFeaturesPath'])
writeSheet = writeBook.active

for i in range(writeSheet.max_row - 1):
    punctuation = writeSheet.cell(row = i + 2, column = 7).value
    tag = 0
    if punctuation in configs['punctuation']:
        tag = configs['punctuation'][punctuation]
    writeSheet['G' + str(i + 2)] = tag
    print('index[', i + 2, ']', punctuation, ',tag=', tag)

writeBook.save(configs['extractFeaturesPath'])