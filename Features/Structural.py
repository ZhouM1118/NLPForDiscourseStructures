import openpyxl
import nltk
import re
import config
configs = config.configs
import os
from nltk.parse import stanford
from datetime import datetime
# from nltk.tokenize import sent_tokenize

os.environ["STANFORD_PARSER"] = configs['stanfordParserPath']
os.environ["STANFORD_MODELS"] = configs['stanfordParserModelsPath']
parser = stanford.StanfordParser(model_path=u"edu/stanford/nlp/models/lexparser/englishPCFG.ser.gz")

nltk.data.path.append(configs['nltkDataPath'])

# 记录程序开始执行时间
start = datetime.now()

readBook = openpyxl.load_workbook(configs['dataSetPath'])
readSheet = readBook.active

writeBook = openpyxl.load_workbook(configs['extractFeaturesPath'])
writeSheet = writeBook.active

# 分割不带标签的句子 注：无法识别标点符号后不带空格的句子，故放弃这种方式
# def segregateSentence(paraContent):
#     # sent_tokenizer = nltk.data.load('tokenizers/punkt/english.pickle')
#     # sentences = sent_tokenizer.tokenize(paraContent)
#     sentences = sent_tokenize(paraContent)
#     return sentences

# 分割带标签的句子
def segregateSentenceByTag(paraContentTag):
    sentencesTag = re.split(r'(<[A-Z]{2,4}>)', paraContentTag)
    tags = []
    sentences = []

    for index in range(len(sentencesTag)):
        if sentencesTag[index] == "":
            continue
        elif index%2 == 0:
            sentence = re.split(r'(</[A-Z]{2,4}>)', sentencesTag[index])
            sentences.append(sentence[0].strip())
            continue
        else:
            tags.append(configs['Tags'][sentencesTag[index]])

    return sentences, tags

# 获取一个句子中的单词个数以及句子的标点符号
def getWordCountAndPunctuation(sentence):
    words = re.split(r"\s+", sentence)
    return  len(words), words[-1][-1]
    #return len(nltk.word_tokenize(sentence))

# 获取一个句子的分析树的深度
def getParseTreeDepth(sentence):
    trees = parser.raw_parse(sentence)
    tree = next(trees)
    if tree is None:
        return 0
    return tree.height()

# 获取句子时态特征
def getSentenceTense(sentence):
    #分词
    tokens = nltk.word_tokenize(sentence)
    #词性标注
    tags = nltk.pos_tag(tokens)
    i_1 = 0
    i_2 = 0
    for tag in tags:
        if tag[1] not in configs['tense']:
            continue
        elif configs['tense'][tag[1]] == 1:
            i_1 += 1
        elif configs['tense'][tag[1]] == 2:
            i_2 += 1
    return 1 if i_1 >= i_2 else 2

rows = []

for i in range(readSheet.max_row - 51):
    paraContentTag = readSheet.cell(row = i + 2, column = 8).value
    sentencesAndTag = segregateSentenceByTag(paraContentTag)

    sentences = sentencesAndTag[0]
    tags = sentencesAndTag[1]

    for index in range(len(sentences)):
        print((i + 2),sentences[index])
        wordCountAndPunctuation = getWordCountAndPunctuation(sentences[index])

        row = [readSheet.cell(row = i + 2, column = 1).value,#ID
               readSheet.cell(row = i + 2, column = 2).value,#ParaType
               readSheet.cell(row = i + 2, column = 5).value,#Structure
               sentences[index],#SentenceContent
               tags[index],#SentenceTag
               wordCountAndPunctuation[0],#WordCount
               wordCountAndPunctuation[1],#Punctuation
               index,#position
               getParseTreeDepth(sentences[index]),#parseTreeDepth
               getSentenceTense(sentences[index]),#tense
               '0%'
               ]

        for indicator in configs['indicators']:
            row.append(1) if indicator in sentences[index] \
                             or indicator.capitalize() in sentences[index] else row.append(0)

        rows.append(row)


for row in rows:
    writeSheet.append(row)

writeBook.save(configs['extractFeaturesPath'])

# 测试代码
# paraContentTag = readSheet.cell(row = 68, column = 8).value
# tags = segregateSentenceTag(paraContentTag)
# print(tags)
# print(len(tags))
#
# paraContent = readSheet.cell(row = 68, column = 4).value
# sentences = segregateSentence(paraContent)
# print(sentences)
# print(len(sentences))
#
# from nltk.tokenize import sent_tokenize
# sents = sent_tokenize(readSheet.cell(row = 68, column = 4).value)
# print(sents)
# print(len(sents))
#
# from nltk.tokenize import PunktSentenceTokenizer
# punkts = PunktSentenceTokenizer().tokenize(readSheet.cell(row = 68, column = 4).value)
# print(punkts)
# print(len(punkts))

# segregateSentenceTag(readSheet.cell(row = 68, column = 8).value)
#
# for s in sentences:
#     print(s)

# for index in range(len(sentences)):
#     print(index)
#     print(tags[index])

# 计算程序运行总时间(秒)
elapsed = (datetime.now() - start).seconds
print('Time used : ', elapsed)