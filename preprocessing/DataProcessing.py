import openpyxl
import nltk
import re
import config
import os
from nltk.parse import stanford
from datetime import datetime
# from nltk.tokenize import sent_tokenize

configs = config.configs
os.environ["STANFORD_PARSER"] = configs['stanfordParserPath']
os.environ["STANFORD_MODELS"] = configs['stanfordParserModelsPath']
nltk.data.path.append(configs['nltkDataPath'])

data_set_path = configs['dataSetPath']
data2_set_path = configs['data2SetPath']
dataset_new_score_122_path = configs['dataset_newScore_122_path']
new_score_122_path = configs['dataset_newScore_122_path']
new_scores_path = configs['new_scores_path']
features_path = configs['featuresPath']
condensed_features_path = configs['condensedFeaturesPath']
test_features_path = configs['testFeaturesPath']
test2_path = configs['test2Path']
test3_path = configs['test3Path']
test4_path = configs['test4Path']
test5_path = configs['test5Path']
test_last_20_add20_Path = configs['test_last_20_add20_Path']
test_last_20_add30_Path = configs['test_last_20_add30_Path']
test_last_20_add40_Path = configs['test_last_20_add40_Path']
test_last_20_allFeatures_Path = configs['test_last_20_allFeatures_Path']
condensed_test_features_path = configs['condensedTestFeaturesPath']
all_features_path = configs['allFeaturesPath']
all_features_add20_path = configs['allFeatures_add20_Path']
all_features_add30_path = configs['allFeatures_add30_Path']
all_features_add40_path = configs['allFeatures_add40_Path']
parser = stanford.StanfordParser(model_path=u"edu/stanford/nlp/models/lexparser/englishPCFG.ser.gz")


def time(func):
    def wrapper(*args):
        # 记录程序开始执行时间
        start = datetime.now()
        result = func(*args)
        elapsed = (datetime.now() - start).seconds
        print('"', func.__name__, '"function :', 'Time used : ', elapsed, 's')
        return result
    return wrapper


class DataProcessing(object):
    """数据预处理类，提供所有的预处理操作"""

    @staticmethod
    def segregate_sentence(para_content):
        """
        分割不带标签的句子 注：无法识别标点符号后不带空格的句子，故放弃这种方式
        :param para_content: 段落内容
        :return: 段落中的句子列表
        """
        sent_tokenizer = nltk.data.load('tokenizers/punkt/english.pickle')
        sentences = sent_tokenizer.tokenize(para_content)
        return sentences

    @staticmethod
    def segregate_sentence_by_tag(para_content_tag):
        """
        分割带标签的段落
        :param para_content_tag: 代表标签的段落内容
        :return:
            sentences:句子列表
            tags:句子标签
        """
        sentences_tag = re.split(r'(<[A-Z]{2,4}>)', para_content_tag)
        tags = []
        sentences = []

        for index in range(len(sentences_tag)):
            if sentences_tag[index] == "":
                continue
            elif index % 2 == 0:
                sentence = re.split(r'(</[A-Z]{2,4}>)', sentences_tag[index])
                sentences.append(sentence[0].strip())
                continue
            else:
                tags.append(configs['Tags'][sentences_tag[index]])

        return sentences, tags

    @staticmethod
    def get_wordcount_and_punctuation(sentence):
        """
        获取一个句子中的单词个数以及标点符号
        :param sentence: 句子
        :return: 句子单词个数，句子标签
        """

        words = re.split(r"\s+", sentence)
        punctuation = words[-1][-1]
        tag = 0
        if punctuation in configs['punctuation']:
            tag = configs['punctuation'][punctuation]
        return len(words), tag

    @staticmethod
    def get_parsetree_depth_and_subclause_num(sentence):
        """
        获取一个句子的分析树的深度以及句子的子句数
        :param sentence: 句子
        :return:
            height：分析树的深度
            subClauseNum：子句数
        """
        trees = parser.raw_parse(sentence)
        tree = next(trees)
        height = 0
        subclause_num = 0
        if tree:
            height = tree.height()
            for subtree in tree.subtrees():
                if subtree.label() == 'SBAR':
                    subclause_num += 1

        return height, subclause_num

    @staticmethod
    def get_sentence_tense_and_pos_flag(sentence):
        """
        获取句子时态特征以及词性标注的各种特征
        :param sentence: 句子
        :return:
            tense_flag：时态特征，
            NNP_flag：专有名词特征,
            MD_flag：情态动词特征,
            PRP_flag：人称代词特征
        """
        # 分词
        tokens = nltk.word_tokenize(sentence)
        # 词性标注
        tags = nltk.pos_tag(tokens)
        NNP_flag = 0
        MD_flag = 0
        PRP_flag = 0
        i_1 = 0
        i_2 = 0
        for tag in tags:
            if tag[1] == 'NNP':
                NNP_flag = 1
            if tag[1] == 'MD':
                MD_flag = 1
            if tag[1] == 'PRP' or tag[1] == 'PRP$':
                PRP_flag += 1
            if tag[1] not in configs['tense']:
                continue
            elif configs['tense'][tag[1]] == 1:
                i_1 += 1
            elif configs['tense'][tag[1]] == 2:
                i_2 += 1
        tense_flag = 1 if i_1 >= i_2 else 2
        return tense_flag, NNP_flag, MD_flag, PRP_flag

    @staticmethod
    @time
    def extract_train_structural_feature():
        """
        提取训练集中的篇章结构特征并持久化
        :return:
        """

        read_book = openpyxl.load_workbook(data_set_path)
        read_sheet = read_book.active
        write_book = openpyxl.load_workbook(features_path)
        write_sheet = write_book.active

        rows = []

        for i in range(read_sheet.max_row - 51):
            para_content_tag = read_sheet.cell(row=i + 2, column=8).value
            sentences_and_tag = DataProcessing.segregate_sentence_by_tag(para_content_tag)

            sentences = sentences_and_tag[0]
            tags = sentences_and_tag[1]

            for index in range(len(sentences)):
                print((i + 2), sentences[index])
                word_count_and_punctuation = DataProcessing.get_wordcount_and_punctuation(sentences[index])
                tense_and_pos_flag = DataProcessing.get_sentence_tense_and_pos_flag(sentences[index])
                parsetree_depth_and_subclause_num = DataProcessing.get_parsetree_depth_and_subclause_num(sentences[index])

                row = [read_sheet['A' + str(i + 2)].value,  # ID
                       read_sheet['B' + str(i + 2)].value,  # ParaType
                       read_sheet['E' + str(i + 2)].value,  # Structure
                       sentences[index],  # SentenceContent
                       tags[index],  # SentenceTag
                       configs['paraType'][read_sheet['B' + str(i + 2)].value],  # ParaTag
                       0, 0,  # BeforeSentenceTag,AfterSentenceTag
                       word_count_and_punctuation[0],  # WordCount
                       word_count_and_punctuation[1],  # Punctuation
                       index,  # position
                       parsetree_depth_and_subclause_num[0],  # parseTreeDepth
                       tense_and_pos_flag[0],  # tense
                       tense_and_pos_flag[1],  # NNPFlag
                       tense_and_pos_flag[2],  # MD
                       parsetree_depth_and_subclause_num[1],  # subClauseNum
                       0, 0, 0, 0, 0, 0, 0, 0, 0
                       ]

                for indicator in configs['indicators']:
                    row.append(1) if indicator in sentences[index] \
                                     or indicator.capitalize() in sentences[index] else row.append(0)

                rows.append(row)

        for row in rows:
            write_sheet.append(row)

        write_book.save(features_path)

    @staticmethod
    @time
    def extract_test_structural_feature():
        """
        提取测试集中的篇章结构特征并持久化
        :return:
        """

        read_book = openpyxl.load_workbook(data2_set_path)
        read_sheet = read_book.active
        write_test_book = openpyxl.load_workbook(test5_path)
        write_test_sheet = write_test_book.active

        rows = []

        for i in range(53):
            k = i + 238
            para_content_tag = read_sheet['H' + str(k)].value
            sentences_and_tag = DataProcessing.segregate_sentence_by_tag(para_content_tag.strip())

            sentences = sentences_and_tag[0]
            tags = sentences_and_tag[1]

            for index in range(len(sentences)):
                print(str(k), sentences[index])
                word_count_and_punctuation = DataProcessing.get_wordcount_and_punctuation(sentences[index])
                tense_and_pos_flag = DataProcessing.get_sentence_tense_and_pos_flag(sentences[index])
                parsetree_depth_and_subclause_num = DataProcessing.get_parsetree_depth_and_subclause_num(sentences[index])

                row = [read_sheet['A' + str(k)].value,  # ID
                       read_sheet['B' + str(k)].value,  # ParaType
                       read_sheet['E' + str(k)].value,  # Structure
                       sentences[index],  # SentenceContent
                       tags[index],  # SentenceTag
                       # configs['paraType'][read_sheet['B' + str(k)].value],  # ParaTag
                       0, 0,  # BeforeSentenceTag,AfterSentenceTag
                       word_count_and_punctuation[0],  # WordCount
                       word_count_and_punctuation[1],  # Punctuation
                       index,  # position
                       parsetree_depth_and_subclause_num[0],  # parseTreeDepth
                       tense_and_pos_flag[0],  # tense
                       tense_and_pos_flag[1],  # NNP
                       tense_and_pos_flag[2],  # MD
                       tense_and_pos_flag[3],  # PRP
                       parsetree_depth_and_subclause_num[1],  # subClauseNum
                       0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
                       ]

                for indicator in configs['condensedIndicators']:
                    row.append(1) if indicator in sentences[index] \
                                     or indicator.capitalize() in sentences[index] else row.append(0)

                rows.append(row)

        for row in rows:
            write_test_sheet.append(row)

        write_test_book.save(test5_path)

    @staticmethod
    @time
    def extract_structural_and_score():
        """
        提取所有语料库中的篇章结构分数
        :return:
        """

        read_book = openpyxl.load_workbook(new_score_122_path)
        read_sheet = read_book.active
        write_test_book = openpyxl.load_workbook(new_scores_path)
        write_test_sheet = write_test_book.active

        rows = []

        for i in range(read_sheet.max_row - 1):
            k = i + 2
            para_content_tag = read_sheet['H' + str(k)].value
            sentences_and_tag = DataProcessing.segregate_sentence_by_tag(para_content_tag.strip())

            sentences = sentences_and_tag[0]
            tags = sentences_and_tag[1]

            for index in range(len(sentences)):
                print(str(k), sentences[index])

                row = [read_sheet['A' + str(k)].value,  # ID
                       read_sheet['B' + str(k)].value,  # ParaType
                       read_sheet['E' + str(k)].value,  # Structure
                       sentences[index],  # SentenceContent
                       tags[index],  # SentenceTag
                       ]

                rows.append(row)

        for row in rows:
            write_test_sheet.append(row)

        write_test_book.save(new_scores_path)

    @staticmethod
    def extract_tag_sentence_file():
        """
        从featuresPath中分别提取不同标签的句子到指定路径下的对应文件中
        :return:
        """
        read_book = openpyxl.load_workbook(all_features_add40_path)
        read_sheet = read_book.active

        BG_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_BGFile.txt', 'a')
        PT_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_PTFile.txt', 'a')
        TST_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_TSTFile.txt', 'a')
        RS_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_RSFile.txt', 'a')
        REXP_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_REXPFile.txt', 'a')
        EG_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_EGFile.txt', 'a')
        EEXP_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_EEXPFile.txt', 'a')
        GRL_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_GRLFile.txt', 'a')
        ADM_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_ADMFile.txt', 'a')
        RTT_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_RTTFile.txt', 'a')
        SRS_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_SRSFile.txt', 'a')
        RAFM_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_RAFMFile.txt', 'a')
        IRL_file = open('/Users/ming.zhou/NLP/datasets/ngram/test2_add40_IRLFile.txt', 'a')

        BG_content = []
        PT_content = []
        TST_content = []
        RS_content = []
        REXP_content = []
        EG_content = []
        EEXP_content = []
        GRL_content = []
        ADM_content = []
        RTT_content = []
        SRS_content = []
        RAFM_content = []
        IRL_content = []

        put_sentence_by_tag = {
            1: lambda x: BG_content.append(x),
            2: lambda x: PT_content.append(x),
            3: lambda x: TST_content.append(x),
            4: lambda x: RS_content.append(x),
            5: lambda x: REXP_content.append(x),
            6: lambda x: EG_content.append(x),
            7: lambda x: EEXP_content.append(x),
            8: lambda x: GRL_content.append(x),
            9: lambda x: ADM_content.append(x),
            10: lambda x: RTT_content.append(x),
            11: lambda x: SRS_content.append(x),
            12: lambda x: RAFM_content.append(x),
            13: lambda x: IRL_content.append(x)
        }

        for i in range(read_sheet.max_row - 1):
            sentence_tag = read_sheet['E' + str(i + 2)].value
            sentence = read_sheet['D' + str(i + 2)].value
            put_sentence_by_tag[sentence_tag](sentence + '\n')

        BG_file.writelines(BG_content)
        BG_file.close()
        PT_file.writelines(PT_content)
        PT_file.close()
        TST_file.writelines(TST_content)
        TST_file.close()
        RS_file.writelines(RS_content)
        RS_file.close()
        REXP_file.writelines(REXP_content)
        REXP_file.close()
        EG_file.writelines(EG_content)
        EG_file.close()
        EEXP_file.writelines(EEXP_content)
        EEXP_file.close()
        GRL_file.writelines(GRL_content)
        GRL_file.close()
        ADM_file.writelines(ADM_content)
        ADM_file.close()
        RTT_file.writelines(RTT_content)
        RTT_file.close()
        SRS_file.writelines(SRS_content)
        SRS_file.close()
        RAFM_file.writelines(RAFM_content)
        RAFM_file.close()
        IRL_file.writelines(IRL_content)
        IRL_file.close()

    @staticmethod
    def extract_all_sentence(from_file_path, to_file_path):
        """
        从指定路径（fromFilePath）下读取文件内容并将其中所有的句子提取到指定路径下（toFilePath）
        :param from_file_path: 读取文件路径
        :param to_file_path: 目标路径
        :return:
        """
        read_book = openpyxl.load_workbook(from_file_path)
        read_sheet = read_book.active

        all_sentence_file = open(to_file_path, 'a')
        sentences = []

        for i in range(read_sheet.max_row - 1):
            sentence = read_sheet['D' + str(i + 2)].value
            sentences.append(sentence + '\n')

        all_sentence_file.writelines(sentences)
        all_sentence_file.close()

    @staticmethod
    def integrate_sentencetag_context_and_paratag_feature():
        """
        设置特征文件中句子的上一句和下一句的句子标签特征以及段落标签特征
        :return:
        """

        write_book = openpyxl.load_workbook(test_last_20_allFeatures_Path)
        write_sheet = write_book.active

        before_essay_id = 0
        for i in range(write_sheet.max_row - 1):
            # 设置特征文件中句子的段落标签特征
            # para_type = write_sheet['B' + str(i + 2)].value
            # write_sheet['F' + str(i + 2)] = configs['para_type'][para_type]

            now_essay_id = write_sheet['A' + str(i + 2)].value.strip().split('-')[0]
            is_new_essay = now_essay_id != before_essay_id

            if is_new_essay:
                write_sheet['F' + str(i + 2)] = 0  # 设置当前句的BeforeSentenceTag为0
                if i > 0:
                    write_sheet['G' + str(i + 1)] = 0  # 设置上一句的AfterSentenceTag为0
                write_sheet['G' + str(i + 2)] = write_sheet['E' + str(i + 3)].value * 10  # 设置当前句的AfterSentenceTag
            else:
                write_sheet['F' + str(i + 2)] = write_sheet['E' + str(i + 1)].value * 10  # 设置当前句的BeforeSentenceTag
                if (i + 2) == write_sheet.max_row:
                    write_sheet['G' + str(i + 2)] = 0
                else:
                    write_sheet['G' + str(i + 2)] = write_sheet['E' + str(i + 3)].value * 10  # 设置当前句的AfterSentenceTag

            before_essay_id = write_sheet['A' + str(i + 2)].value.strip().split('-')[0]
        write_book.save(test_last_20_allFeatures_Path)

    @staticmethod
    def integrate_ngram_feature(ngram_result_file_path, feature_file_column_index):
        """
        设置特征文件中句子的NGram特征
        :param ngram_result_file_path: ngram特征文件路径
        :param feature_file_column_index: 特征文件中指定ngram特征的列标号
        :return:
        """
        write_test_book = openpyxl.load_workbook(test_last_20_add40_Path)
        write_test_sheet = write_test_book.active

        result_file = open(ngram_result_file_path)

        i = 2
        for line in result_file:
            result = re.findall('ppl=.*? ppl', line)
            if result:
                write_test_sheet[feature_file_column_index + str(i)] = re.split('=', result[0])[1].strip().split(' ')[0]
                i += 1

        write_test_book.save(test_last_20_add40_Path)

    @staticmethod
    def add_new_indicators(file_path, indicators, indicator_index):
        """
        添加指示词特征维
        :param file_path: 特征文件路径
        :param indicators: 指示词列表
        :param indicator_index: 指示词列标号
        :return:
        """
        write_book = openpyxl.load_workbook(file_path)
        write_sheet = write_book.active

        for index in range(write_sheet.max_row - 1):
            sentence = write_sheet['D' + str(index + 2)].value
            row = []
            for i in range(len(indicators)):
                if indicators[i] in sentence or indicators[i].capitalize() in sentence:
                    row.append(1)
                else:
                    row.append(0)

            for k in range(len(indicator_index)):
                write_sheet[indicator_index[k] + str(index + 2)] = row[k]

        write_book.save(file_path)

    @staticmethod
    def add_feature_column(file_path, feature_name, feature_index):
        """
        添加指定特征的特征维
        :param file_path: 特征文件路径
        :param feature_name: 特征名称
        :param feature_index: 特征的列标号
        :return:
        """
        write_book = openpyxl.load_workbook(file_path)
        write_sheet = write_book.active

        # 设置特征列名
        write_sheet[feature_index + str(1)] = feature_name
        for index in range(write_sheet.max_row - 1):
            sentence = write_sheet['D' + str(index + 2)].value
            write_sheet[feature_index + str(index + 2)] = DataProcessing.get_sentence_tense_and_pos_flag(sentence)[3]

        write_book.save(file_path)


# indicators = ['above', 'conclusion', 'agree', 'admittedly']
# indicator_index = ['CF', 'CG', 'CH', 'CI']
# DataProcessing.addFeatureColumn('/Users/ming.zhou/NLP/datasets/eliminateParaTagTest.xlsx', 'PRP', 'P')
# DataProcessing.addNewIndicators(condensedTestFeaturesPath, indicators, indicatorIndex)
# DataProcessing.extract_test_structural_feature()
fromPath = '/Users/ming.zhou/NLP/datasets/allFeatures_add40.xlsx'
toPath = '/Users/ming.zhou/NLP/datasets/ngram/test/allFeatures_add40_File.txt'
# DataProcessing.extract_all_sentence(fromPath, toPath)
ngramPath = '/Users/ming.zhou/NLP/datasets/ngram/modal/TEST_Test_last_20_add40_RAFMResult'
# DataProcessing.integrate_ngram_feature(ngramPath, 'AA')
# DataProcessing.integrate_sentencetag_context_and_paratag_feature()
# DataProcessing.extractTagSentenceFile()
# DataProcessing.extract_tag_sentence_file()
# DataProcessing.extract_structural_and_score()
