import openpyxl
import config
configs = config.configs

readBook = openpyxl.load_workbook(configs['extractFeaturesPath'])
# 获取当前正在显示的sheet
readSheet = readBook.active

BGFile = open('/Users/ming.zhou/NLP/datasets/BGFile.txt', 'a')
PTFile = open('/Users/ming.zhou/NLP/datasets/PTFile.txt', 'a')
TSTFile = open('/Users/ming.zhou/NLP/datasets/TSTFile.txt', 'a')
RSFile = open('/Users/ming.zhou/NLP/datasets/RSFile.txt', 'a')
REXPFile = open('/Users/ming.zhou/NLP/datasets/REXPFile.txt', 'a')
EGFile = open('/Users/ming.zhou/NLP/datasets/EGFile.txt', 'a')
EEXPFile = open('/Users/ming.zhou/NLP/datasets/EEXPFile.txt', 'a')
GRLFile = open('/Users/ming.zhou/NLP/datasets/GRLFile.txt', 'a')
ADMFile = open('/Users/ming.zhou/NLP/datasets/ADMFile.txt', 'a')
RTTFile = open('/Users/ming.zhou/NLP/datasets/RTTile.txt', 'a')
SRSFile = open('/Users/ming.zhou/NLP/datasets/SRSFile.txt', 'a')
RAFMFile = open('/Users/ming.zhou/NLP/datasets/RAFMFile.txt', 'a')
IRLFile = open('/Users/ming.zhou/NLP/datasets/IRLFile.txt', 'a')

BGContent = []
PTContent = []
TSTContent = []
RSContent = []
REXPContent = []
EGContent = []
EEXPContent = []
GRLContent = []
ADMContent = []
RTTContent = []
SRSContent = []
RAFMContent = []
IRLContent = []

putSentenceByTag = {
    1 : lambda x: BGContent.append(x) ,
    2: lambda x: PTContent.append(x),
    3: lambda x: TSTContent.append(x),
    4: lambda x: RSContent.append(x),
    5: lambda x: REXPContent.append(x),
    6: lambda x: EGContent.append(x),
    7: lambda x: EEXPContent.append(x),
    8: lambda x: GRLContent.append(x),
    9: lambda x: ADMContent.append(x),
    10: lambda x: RTTContent.append(x),
    11: lambda x: SRSContent.append(x),
    12: lambda x: RAFMContent.append(x),
    13: lambda x: IRLContent.append(x)
}

for i in range(readSheet.max_row - 1):
    sentenceTag = readSheet.cell(row = i + 2, column = 5).value
    sentence = readSheet.cell(row = i + 2, column = 4).value
    putSentenceByTag[sentenceTag](sentence  + '\n')

BGFile.writelines(BGContent)
BGFile.close()
PTFile.writelines(PTContent)
PTFile.close()
TSTFile.writelines(TSTContent)
TSTFile.close()
RSFile.writelines(RSContent)
RSFile.close()
REXPFile.writelines(REXPContent)
REXPFile.close()
EGFile.writelines(EGContent)
EGFile.close()
EEXPFile.writelines(EEXPContent)
EEXPFile.close()
GRLFile.writelines(GRLContent)
GRLFile.close()
ADMFile.writelines(ADMContent)
ADMFile.close()
RTTFile.writelines(RTTContent)
RTTFile.close()
SRSFile.writelines(SRSContent)
SRSFile.close()
RAFMFile.writelines(RAFMContent)
RAFMFile.close()
IRLFile.writelines(IRLContent)
IRLFile.close()