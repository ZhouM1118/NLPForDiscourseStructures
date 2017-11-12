import openpyxl
import config
configs = config.configs

writeBook = openpyxl.load_workbook(configs['extractFeaturesPath'])
writeSheet = writeBook.active

writeTestBook = openpyxl.load_workbook(configs['extractTestFeaturesPath'])
writeTestSheet = writeTestBook.active

import re
def integrateNGramFeature():

    # BGResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/BGResult')
    # PTResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/PTResult')
    # REXPResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/REXPResult')
    # EGResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/EGResult')
    # EEXPResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/EEXPResult')
    # ADMResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/ADMResult')
    # RTTResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/RTTResult')
    # SRSResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/SRSResult')
    # RAFMResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/RAFMResult')
    # TestBGResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/TestBGResult')
    # TestPTResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/TestPTResult')
    # TestREXPResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/TestREXPResult')
    # TestEGResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/TestEGResult')
    # TestEEXPResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/TestEEXPResult')
    # TestADMResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/TestADMResult')
    # TestRTTResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/TestRTTResult')
    # TestSRSResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/TestSRSResult')
    TestRAFMResultFile = open('/Users/ming.zhou/NLP/datasets/ngram/modal/TestRAFMResult')

    i = 2
    for line in TestRAFMResultFile:
        result = re.findall('ppl=.*? ppl', line)
        if result:
            writeTestSheet['V' + str(i)] = re.split('=', result[0])[1].strip().split(' ')[0]
            i += 1
            # if 'zeroprobs' in line and 'logprob=' in line and 'ppl=' in line and 'ppl1=' in line:
            #     print(line)
    writeTestBook.save(configs['extractTestFeaturesPath'])

def integrateSentenceTagFeature():
    nowEssayId = 0
    beforeEssayId = 0
    for i in range(writeSheet.max_row - 1):
        nowEssayId = writeSheet['A' + str(i + 2)].value.strip().split('-')[0]
        isNewEssay = nowEssayId != beforeEssayId

        if isNewEssay:
            writeSheet['G' + str(i + 2)] = 0 #设置当前句的BeforeSentenceTag为0
            if i > 0:
                writeSheet['H' + str(i + 1)] = 0 #设置上一句的AfterSentenceTag为0
            writeSheet['H' + str(i + 2)] = writeSheet['E' + str(i + 3)].value * 10 #设置当前句的AfterSentenceTag
        else:
            writeSheet['G' + str(i + 2)] = writeSheet['E' + str(i + 1)].value * 10 #设置当前句的BeforeSentenceTag
            if (i + 2) == writeSheet.max_row:
                writeSheet['H' + str(i + 2)] = 0
            else:
                writeSheet['H' + str(i + 2)] = writeSheet['E' + str(i + 3)].value * 10 #设置当前句的AfterSentenceTag

        beforeEssayId = writeSheet['A' + str(i + 2)].value.strip().split('-')[0]
    writeBook.save(configs['extractFeaturesPath'])

def integrateParaTagFeature():
    for i in range(writeSheet.max_row - 1):
        paraType = writeSheet['B' + str(i + 2)].value
        writeSheet['F' + str(i + 2)] = configs['paraType'][paraType]
    writeBook.save(configs['extractFeaturesPath'])

def integratePunctuationFeature():
    for i in range(writeTestSheet.max_row - 1):
        punctuation = writeTestSheet['J' + str(i + 2)].value
        tag = 0
        if punctuation in configs['punctuation']:
            tag = configs['punctuation'][punctuation]
        writeTestSheet['J' + str(i + 2)] = tag
        print('index[', i + 2, ']', punctuation, ',tag=', tag)

    writeTestBook.save(configs['extractTestFeaturesPath'])

integrateNGramFeature()