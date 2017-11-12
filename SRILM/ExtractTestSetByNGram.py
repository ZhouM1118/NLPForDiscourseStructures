import openpyxl
import config
configs = config.configs

# 提取在训练集中所有的句子
def extractTestContent():

    readBook = openpyxl.load_workbook(configs['extractFeaturesPath'])
    # 获取当前正在显示的sheet
    readSheet = readBook.active

    testFile = open('/Users/ming.zhou/NLP/datasets/ngram/test/testFile.txt', 'a')
    testContent = []

    for i in range(readSheet.max_row - 1):
        sentence = readSheet.cell(row=i + 2, column=4).value
        testContent.append(sentence + '\n')

    testFile.writelines(testContent)
    testFile.close()

# 提取测试模型的句子
def extractTestSetContent():

    readTestBook = openpyxl.load_workbook(configs['extractTestFeaturesPath'])
    # 获取当前正在显示的sheet
    readTestSheet = readTestBook.active

    testSetFile = open('/Users/ming.zhou/NLP/datasets/ngram/test/testSetFile.txt', 'a')
    testSetContent = []

    for i in range(readTestSheet.max_row - 1):
        sentence = readTestSheet['D' + str(i + 2)].value
        testSetContent.append(sentence + '\n')

    testSetFile.writelines(testSetContent)
    testSetFile.close()

extractTestSetContent()