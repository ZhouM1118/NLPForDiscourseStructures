import openpyxl
import config
import math
import numpy
from sklearn import svm
from sklearn.ensemble import RandomForestClassifier
from sklearn.ensemble import ExtraTreesClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.metrics import accuracy_score
from sklearn.model_selection import train_test_split
from sklearn import cross_validation
from sklearn.metrics import roc_curve, auc
from sklearn import preprocessing
from sklearn.grid_search import GridSearchCV
# from sklearn.cross_validation import cross_val_score
from sklearn.metrics import hamming_loss
import pylab as pl
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from datetime import datetime

configs = config.configs

# 记录程序开始执行时间
startTime = datetime.now()

# readBook = openpyxl.load_workbook(configs['condensedFeaturesPath'])
# readSheet = readBook.active

# read_book = openpyxl.load_workbook('/Users/ming.zhou/NLP/datasets/eliminateParaTag.xlsx')
# read_sheet = read_book.active

# readTestBook = openpyxl.load_workbook(configs['condensedTestFeaturesPath'])
# readTestSheet = readTestBook.active

# read_test_book = openpyxl.load_workbook('/Users/ming.zhou/NLP/datasets/eliminateParaTagTest.xlsx')
# read_test_sheet = read_test_book.active

read_all_book = openpyxl.load_workbook(configs['allFeatures_add40_Path'])
read_all_sheet = read_all_book.active

read_test2_book = openpyxl.load_workbook(configs['test_last_20_add40_Path'])
read_test2_sheet = read_test2_book.active


def get_feature_vector_and_tag():
    """
    获取训练集的句子特征向量以及句子标签
    :return:
        feature_vector_list：句子特征向量列表
        sentence_tag：句子标签
    """

    feature_vector_list = []
    sentence_tag = []
    for i in range(read_all_sheet.max_row - 1):

        feature_vector = []
        sentence_tag.append(read_all_sheet['E' + str(i + 2)].value)
        # 先添加段落特征
        feature_vector.append(float(read_all_sheet['A' + str(i + 2)].value.strip().split('-')[1]))
        for j in range(read_all_sheet.max_column - 5):
            if isinstance(read_all_sheet.cell(row=i + 2, column=j + 6).value, str):
                feature_vector.append(float(read_all_sheet.cell(row=i + 2, column=j + 6).value))
            else:
                feature_vector.append(read_all_sheet.cell(row=i + 2, column=j + 6).value)
        feature_vector_list.append(feature_vector)
    return feature_vector_list, sentence_tag


def get_tag_and_score(file_path, start, end, para_type='Introduction'):
    """
    获取训练集的句子标签以及段落分数
    :return:
    """
    read_book = openpyxl.load_workbook(file_path)
    read_sheet = read_book.active
    tags = []
    tag = []
    scores = []
    now_id = read_sheet['A2'].value
    sentence_tags = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]
    max_length = 0

    for i in range(start, end):
        essay_id = read_sheet['A' + str(i)].value
        para_type_value = read_sheet['B' + str(i)].value
        #  判断是否是新的段落
        if essay_id != now_id:
            now_id = essay_id
            if len(tag) != 0:
                tags.append(tag)
                score = read_sheet['C' + str(i - 1)].value
                scores.append(score)
                # 保存最大的tag长度值
                if len(tag) > max_length:
                    max_length = len(tag)
                tag = []
        if para_type == para_type_value:
            tag.append(read_sheet['E' + str(i)].value)
    tag_features = []
    for tag in tags:
        tag_feature = []
        for sentence_tag in sentence_tags:
            sum = 0
            for t in tag:
                if sentence_tag == t:
                    sum += 1
            tag_feature.append(sum)
        tag_features.append(tag_feature)
    # 填充tag数据
    for tag in tags:
        if len(tag) < max_length:
            for j in range(max_length - len(tag)):
                tag.append(0)
    return tags, scores, tag_features


def get_test_feature_vector_and_tag():
    """
    获取测试集的句子特征向量以及句子标签
    :return:
        feature_vector_list：句子特征向量列表
        sentence_tag：句子标签
    """
    feature_vector_list = []
    sentence_tag = []
    for i in range(read_test2_sheet.max_row - 1):

        feature_vector = []
        sentence_tag.append(read_test2_sheet['E' + str(i + 2)].value)
        # 先添加段落特征
        feature_vector.append(float(read_test2_sheet['A' + str(i + 2)].value.strip().split('-')[1]))
        for j in range(read_test2_sheet.max_column - 5):
            if isinstance(read_test2_sheet.cell(row=i + 2, column=j + 6).value, str):
                feature_vector.append(float(read_test2_sheet.cell(row=i + 2, column=j + 6).value))
            else:
                feature_vector.append(read_test2_sheet.cell(row=i + 2, column=j + 6).value)

        # print(len(feature_vector))
        feature_vector_list.append(feature_vector)
    return feature_vector_list, sentence_tag


def adjust_parameter(x, y, flag='ExtraTrees'):
    """
    为模型调参
    :param x: 样本集的特征向量
    :param y: 样本集的标签列表
    :param flag: 模型指示词
    :return:
    """

    clf = ExtraTreesClassifier(min_samples_split=100, min_samples_leaf=20,
                               max_depth=8, max_features='sqrt', random_state=10)
    if flag == 'RandomForest':
        clf = RandomForestClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()
    param_grid = {
        'n_estimators': range(5, 20, 2)
    }
    grid_search = GridSearchCV(estimator=clf, param_grid=param_grid, scoring='roc_auc', cv=5)
    grid_search.fit(x, y)
    print(grid_search.grid_scores_)
    print(grid_search.best_params_)
    print(grid_search.best_score_)


def do_train(x, y, flag='RandomForest'):
    """
    训练并测试训练集的准确度
    :param x: 样本集的特征向量
    :param y: 样本集的标签列表
    :param flag: 模型指示词
    :return:
    """

    clf = RandomForestClassifier(n_estimators=10)
    if flag == 'ExtraTrees':
        clf = ExtraTreesClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    clf.fit(x, y)
    # clf = clf.fit(X, Y)

    y_pred = clf.predict(x)
    print(accuracy_score(y, y_pred))
    print(y)
    print(y_pred)
    print('Y len is %s，y_pred len is %s \n' % (len(y), len(y_pred)))
    # scores = cross_val_score(clf, X, Y)
    # print(scores.mean())

    compare_result = {}
    j = 0
    for i in range(len(y)):
        if y[i] == 5:
            j += 1
        y_and_ypred = str(y[i]) + '-' + str(y_pred[i])
        # print(y_and_ypred)
        if y[i] != y_pred[i]:
            if y_and_ypred not in compare_result:
                compare_result[y_and_ypred] = 1
            else:
                compare_result[y_and_ypred] = compare_result[y_and_ypred] + 1

    print(sorted(compare_result.items(), key=lambda d: -d[1]))


def do_train_by_test_set(train_x, train_y, test_x, test_y, flag='RandomForest'):
    """
    测试测试集的准确度
    :param train_x: 样本集中的特征向量列表
    :param train_y: 样本集中的标签列表
    :param test_x: 测试集中的特征向量列表
    :param test_y: 测试集中的标签列表
    :param flag: 模型指示词
    :return:
    """

    clf = RandomForestClassifier(n_estimators=10)
    file_path = '/Users/ming.zhou/NLP/DiscourseStructures/result/' + flag + 'Result20171115.text'

    if flag == 'ExtraTrees':
        clf = ExtraTreesClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    clf.fit(train_x, train_y)
    y_pred = clf.predict(test_x)

    result = open(file_path, 'a')
    result_content = list()
    result_content.append(str(clf.score(test_x, test_y)) + '\n')
    result_content.append('test_Y:' + '\n')
    result_content.append(str(test_y) + '\n')
    result_content.append('y_pred:' + '\n')
    result_content.append(str(y_pred) + '\n')
    result_content.append('test_Y len is ' + str(len(test_y)) + 'y_pred len is ' + str(len(y_pred)) + '\n')

    print(clf.score(test_x, test_y))
    print(accuracy_score(test_y, y_pred))
    print(test_y)
    print(y_pred)
    print('test_Y len is %s，y_pred len is %s \n' % (len(test_y), len(y_pred)))

    compare_result = {}
    for i in range(len(test_y)):
        y_and_ypred = str(test_y[i]) + '-' + str(y_pred[i])
        # print(y_and_ypred)
        if test_y[i] != y_pred[i]:
            if y_and_ypred not in compare_result:
                compare_result[y_and_ypred] = 1
            else:
                compare_result[y_and_ypred] = compare_result[y_and_ypred] + 1

    sorted_result = sorted(compare_result.items(), key=lambda d: -d[1])

    result_content.append(str(sorted_result) + '\n')
    print(sorted_result)
    result.writelines(result_content)
    result.close()


def do_train_by_cv(x, y, times_num, flag='ExtraTrees'):
    """
    交叉验证模型在训练集上的准确度
    :param x: 样本集中的特征向量列表
    :param y: 样本集中的标签列表
    :param times_num: 模型的n_estimators值
    :param flag: 模型指示词
    :return: 返回模型10次准确度的平均值
    """

    clf = ExtraTreesClassifier(n_estimators=times_num)
    file_path = '/Users/ming.zhou/NLP/DiscourseStructures/result/' + flag + 'CrossValidationResult20171122.text'

    result = open(file_path, 'a')
    result_content = list()
    result_content.append('**********************step=' + str(times_num) + '**********************' + '\n')
    if flag == 'RandomForest':
        clf = RandomForestClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    # 随机划分训练集与测试集，是交叉验证中常用的函数
    sum_num = 0
    for i in range(10):
        x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.1, random_state=i)
        clf.fit(x_train, y_train)
        # Y_pred = clf.predict(x_test)
        score = clf.score(x_test, y_test)
        result_content.append('index[' + str(i) + ']' + str(score) + '\n')
        print(score)

        sum_num += score
    result_content.append('平均准确度：' + str(sum_num/10) + '\n')
    result.writelines(result_content)
    result.close()
    print('平均准确度：', str(sum_num/10))
    return sum_num/10


def do_draw_train_auc(x, y, times_num, tag, flag='ExtraTrees'):
    """
    画出预测模型上的AUC图
    :param x: 样本集中的特征向量列表
    :param y: 样本集中的标签列表
    :param times_num: 模型的n_estimators值
    :param tag: 标签名
    :param flag: 模型指示词
    :return:
    """
    clf = ExtraTreesClassifier(n_estimators=times_num)

    if flag == 'RandomForest':
        clf = RandomForestClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    y_auc = []
    for c in y:
        if configs['Tags'][tag] == c:
            y_auc.append(1)
        else:
            y_auc.append(0)

    x_train, x_test, y_train, y_test = train_test_split(x, y_auc, test_size=0.1, random_state=0)
    clf.fit(x_train, y_train)
    # Y_pred = clf.predict(x_test)
    score = clf.score(x_test, y_test)
    print(score)
    fpr, tpr, threshold = roc_curve(y_test, score)
    roc_auc = auc(fpr, tpr)

    plt.figure()
    lw = 2
    plt.figure(figsize=(10, 10))
    plt.plot(fpr, tpr, color='darkorange',
             lw=lw, label='ROC curve (area = %0.2f)' % roc_auc)  ###假正率为横坐标，真正率为纵坐标做曲线
    plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('Receiver operating characteristic example')
    plt.legend(loc="lower right")
    plt.show()


def do_train_by_cv_and_norm(x, y, times_num, flag='ExtraTrees'):
    """
    使用交叉验证以及规范化技术训练模型在训练集上的准确度
    :param x: 样本集中的特征向量列表
    :param y: 样本集中的标签列表
    :param times_num: 模型的n_estimators值
    :param flag: 模型指示词
    :return:
    """
    clf = ExtraTreesClassifier(n_estimators=10)
    file_path = '/Users/ming.zhou/NLP/DiscourseStructures/result/' + flag + 'CrossValidationResult20171117.text'

    result = open(file_path, 'a')
    result_content = list()
    result_content.append('**********************step=' + str(times_num) + '**********************' + '\n')
    if flag == 'RandomForest':
        clf = RandomForestClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    # 随机划分训练集与测试集，是交叉验证中常用的函数
    sum_num = 0
    for i in range(10):
        x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.1, random_state=i)
        x_train_norm = preprocessing.normalize(x_train, norm='l2')
        x_test_norm = preprocessing.normalize(x_test, norm='l2')
        clf.fit(x_train_norm, y_train)
        # Y_pred = clf.predict(x_test)
        score = clf.score(x_test_norm, y_test)
        result_content.append('index[' + str(i) + ']' + str(score) + '\n')
        print(score)

        sum_num += score
    result_content.append('平均准确度：' + str(sum_num/10) + '\n')
    result.writelines(result_content)
    result.close()
    print('平均准确度：', str(sum_num/10))


def do_train_regression(x, y, times_num, flag='linear'):
    """
    训练评分回归模型
    :param x:
    :param y:
    :return:
    """
    rf = LinearRegression()
    if flag == 'RandomForest':
        rf = RandomForestRegressor(n_estimators=times_num)

    rf.fit(x, y)
    sum_num = 0
    for i in range(10):
        x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.1, random_state=i)
        rf.fit(x_train, y_train)
        # Y_pred = clf.predict(x_test)
        score = rf.score(x_test, y_test)
        sum_num += score
    return sum_num / 10


def do_predict(x_train, y_train, x_test, y_test, times_num, flag='ExtraTrees'):
    """
    预测模型在测试集上的准确度
    :param x_train: 样本集中的特征向量列表
    :param y_train: 样本集中的标签列表
    :param x_test: 测试集中的特征向量列表
    :param y_test: 测试集中的标签列表
    :param times_num: 模型的n_estimators值
    :param flag: 模型指示词
    :return: 返回预测的准确度
    """
    clf = ExtraTreesClassifier(n_estimators=times_num)
    file_path = '/Users/ming.zhou/NLP/DiscourseStructures/result/' + flag + 'PredictResult20171122.text'

    if flag == 'RandomForest':
        clf = RandomForestClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    clf.fit(x_train, y_train)
    y_pred = clf.predict(x_test)
    score = clf.score(x_test, y_test)

    # print(score)
    # print(Y_test)
    # print(y_pred)
    # print('test_Y len is %s，y_pred len is %s \n' %(len(Y_test), len(y_pred)))

    result = open(file_path, 'a')
    result_content = list()
    result_content.append('**********************step=' + str(times_num) + '**********************' + '\n')
    result_content.append(str(score) + '\n')
    result_content.append('test_Y:' + '\n')
    result_content.append(str(y_test) + '\n')
    result_content.append('y_pred:' + '\n')
    result_content.append(str(y_pred) + '\n')
    result_content.append('test_Y len is ' + str(len(y_test)) + '，and y_pred len is ' + str(len(y_pred)) + '\n')

    for i in range(len(y_test)):
        y_and_ypred = str(y_test[i]) + '-' + str(y_pred[i])
        # print(y_and_ypred)
        if y_test[i] != y_pred[i]:
            if y_and_ypred not in compare:
                compare[y_and_ypred] = 1
            else:
                compare[y_and_ypred] = compare[y_and_ypred] + 1

    # sortedResult = sorted(compare.items(), key=lambda d: -d[1])
    # print(sortedResult)
    # result_content.append(str(sortedResult) + '\n')
    result.writelines(result_content)
    result.close()
    return score


def do_draw_predict_auc(x_train, y_train, x_test, y_test, times_num, tag, flag='ExtraTrees'):
    """
    画出预测模型上的AUC图
    :param x_train: 样本集中的特征向量列表
    :param y_train: 样本集中的标签列表
    :param x_test: 测试集中的特征向量列表
    :param y_test: 测试集中的标签列表
    :param times_num: 模型的n_estimators值
    :param tag: 标签名
    :param flag: 模型指示词
    :return:
    """
    y_train_auc = []
    y_test_auc = []
    for tag_train in y_train:
        if configs['Tags'][tag] == tag_train:
            y_train_auc.append(1)
        else:
            y_train_auc.append(0)

    for tag_test in y_test:
        if configs['Tags'][tag] == tag_test:
            y_test_auc.append(1)
        else:
            y_test_auc.append(0)
    plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
    colors = ['magenta', 'green', 'red', 'cyan', 'blue', 'black']
    j = 0
    for i in range(10, 101, 20):
        clf = ExtraTreesClassifier(n_estimators=i)

        if flag == 'RandomForest':
            clf = RandomForestClassifier(n_estimators=10)
        elif flag == 'DecisionTree':
            clf = DecisionTreeClassifier()
        elif flag == 'SVM':
            clf = svm.SVC()

        clf.fit(x_train, y_train_auc)
        y_pred = clf.predict(x_test)
        fpr, tpr, threshold = roc_curve(y_test_auc, y_pred)
        roc_auc = auc(fpr, tpr)
        #  假正率为横坐标，真正率为纵坐标做曲线
        plt.plot(fpr, tpr, color=colors[j],
             lw=2, label='n_estimators %d ROC curve (area = %0.2f)' %(i, roc_auc))
        j += 1
    # y_score = clf.decision_function(x_test)
    # score = clf.score(x_test, y_test_auc)
    # print(y_pred)


    # plt.figure()
    # lw = 2
    # plt.figure(figsize=(10, 10))
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title(tag + 'Receiver operating characteristic')
    plt.legend(loc="lower right")
    plt.show()


def do_predict_by_cv_and_norm(x_train, y_train, x_test, y_test, times_num, flag='ExtraTrees'):
    """
    使用规范化技术预测模型在测试集上的准确度
    :param x_train: 样本集中的特征向量列表
    :param y_train: 样本集中的标签列表
    :param x_test: 测试集中的特征向量列表
    :param y_test: 测试集中的标签列表
    :param times_num: 模型的n_estimators值
    :param flag: 模型指示词
    :return: 返回预测的准确度
    """
    clf = ExtraTreesClassifier(n_estimators=10)
    file_path = '/Users/ming.zhou/NLP/DiscourseStructures/result/' + flag + 'PredictResult20171117.text'

    if flag == 'RandomForest':
        clf = RandomForestClassifier(n_estimators=10)
    elif flag == 'DecisionTree':
        clf = DecisionTreeClassifier()
    elif flag == 'SVM':
        clf = svm.SVC()

    normalizer = preprocessing.Normalizer().fit(x_train)
    x_train_norm = normalizer.transform(x_train)
    x_test_norm = normalizer.transform(x_test)
    clf.fit(x_train_norm, y_train)
    y_pred = clf.predict(x_test_norm)
    score = clf.score(x_test_norm, y_test)

    print(str(score) + '\n')
    print(y_test + '\n')
    print(y_pred + '\n')
    print('test_Y len is %s，y_pred len is %s \n' % (len(y_test), len(y_pred)))

    result = open(file_path, 'a')
    result_content = list()
    result_content.append('**********************step=' + str(times_num) + '**********************' + '\n')
    result_content.append(str(score) + '\n')
    result_content.append('test_Y:' + '\n')
    result_content.append(str(y_test) + '\n')
    result_content.append('y_pred:' + '\n')
    result_content.append(str(y_pred) + '\n')
    result_content.append('test_Y len is ' + str(len(y_test)) + '，and y_pred len is ' + str(len(y_pred)) + '\n')

    for i in range(len(y_test)):
        y_and_ypred = str(y_test[i]) + '-' + str(y_pred[i])

        if y_test[i] != y_pred[i]:
            if y_and_ypred not in compare:
                compare[y_and_ypred] = 1
            else:
                compare[y_and_ypred] = compare[y_and_ypred] + 1

    # sortedResult = sorted(compare.items(), key=lambda d: -d[1])
    # print(sortedResult)
    # result_content.append(str(sortedResult) + '\n')
    result.writelines(result_content)
    result.close()
    return score


def do_draw_train_score(flag):
    """
    画训练集上评分准确率图
    :param flag: 标签名
    :return:
    """
    n = [62, 82, 92, 102, 112, 122]
    scores = []

    tag_and_score_init = get_tag_and_score(configs['new_scores_path'], 2, 1193, flag)
    tag_and_score_add20 = get_tag_and_score(configs['new_scores_path'], 2, 1614, flag)
    tag_and_score_add30 = get_tag_and_score(configs['new_scores_path'], 2, 1824, flag)
    tag_and_score_add40 = get_tag_and_score(configs['new_scores_path'], 2, 2005, flag)
    tag_and_score_add50 = get_tag_and_score(configs['new_scores_path'], 2, 2191, flag)
    tag_and_score_add60 = get_tag_and_score(configs['new_scores_path'], 2, 2414, flag)

    scores.append(do_train_regression(
        tag_and_score_init[2], tag_and_score_init[1], 0))
    scores.append(do_train_regression(
        tag_and_score_add20[2], tag_and_score_add20[1], 0))
    scores.append(do_train_regression(
        tag_and_score_add30[2], tag_and_score_add30[1], 0))
    scores.append(do_train_regression(
        tag_and_score_add40[2], tag_and_score_add40[1], 0))
    scores.append(do_train_regression(
        tag_and_score_add50[2], tag_and_score_add50[1], 0))
    scores.append(do_train_regression(
        tag_and_score_add60[2], tag_and_score_add60[1], 0))

    pl.plot(n, scores, 'g.-')
    pl.title('Linear ' + flag + ' train result')
    pl.xlabel('train set size')
    pl.ylabel('score')
    pl.show()

    # pl.plot(n, score_reason, 'r.-')
    # pl.plot(n, score_concession, 'y.-')
    # pl.plot(n, score_conclusion, 'k.-')
    # # pl.plot(n, score_add50, 'c.-')
    # # pl.plot(n, score_add60, 'b.-')
    # green_patch = mpatches.Patch(color='green', label='introduction score')
    # red_patch = mpatches.Patch(color='red', label='reason score')
    # yellow_patch = mpatches.Patch(color='yellow', label='concession score')
    # black_patch = mpatches.Patch(color='black', label='conclusion score')
    # # cyan_patch = mpatches.Patch(color='cyan', label='add 50 train set')
    # # blue_patch = mpatches.Patch(color='blue', label='add 60 train set')
    # pl.legend(handles=[green_patch, red_patch, yellow_patch, black_patch])
    # pl.title('Linear train result')
    # pl.xlabel('train set size')
    # pl.ylabel('score')
    # pl.show()

params = get_feature_vector_and_tag()
testParams = get_test_feature_vector_and_tag()
# doTrain(params[0], params[1], 'ExtraTrees')
# doTrainByTestSet(params[0], params[1], testParams[0], testParams[1], 'ExtraTrees')
# doTrainByCrossValidation(params[0], params[1], 10, 'ExtraTrees')
# adjustParameter(params[0], params[1])
# doTrainByCVAndNorm(params[0], params[1], 2, 'ExtraTrees')
# print('*********************begin test*********************')
compare = {}


# xs = []
# ys = []
# for index in range(10, 101, 5):
    # xs.append(index)
    # scores = 0
    # for j in range(3):
    #     score = do_predict(params[0], params[1], testParams[0], testParams[1], index, 'ExtraTrees')
    #     scores += score
    # s = do_train_by_cv(params[0], params[1], index, 'ExtraTrees')
    # ys.append(scores/3)
    # ys.append(s)
    # print('index[' + str(index) + ']:' + str(scores/3))
    # print('index[' + str(index) + ']:' + str(s))

# scores = 0
# for j in range(10):
#     score = do_predict(params[0], params[1], testParams[0], testParams[1], 80, 'ExtraTrees')
#     scores += score
# print(scores / 10)

# sortedResult = sorted(compare.items(), key=lambda d: -d[1])
# print(sortedResult)
# print(scores / 30)

# 判断结构评分的数据分布
# y_add60 = get_tag_and_score(configs['allFeatures_add60_Path'], 'Conclusion')[1]
# print(y_add60)
# pl.scatter(y_add60, y_add60, c='green')
# pl.title('Conclusion Scatter Figure')
# pl.xlabel('x')
# pl.ylabel('y')
# pl.legend()
# pl.grid(True)
# pl.show()

# do_draw_train_auc(params[0], params[1], 80, '<EG>')
do_draw_predict_auc(params[0], params[1], testParams[0], testParams[1], 0, '<TST>')
# do_draw_train_score('Reason')
# 计算程序运行总时间(秒)
elapsed = (datetime.now() - startTime).seconds
print('Time used : ', elapsed)
