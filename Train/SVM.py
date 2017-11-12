import openpyxl
import config
import math
from sklearn import svm
from sklearn.metrics import accuracy_score
from sklearn.metrics import hamming_loss
from datetime import datetime

configs = config.configs

# 记录程序开始执行时间
start = datetime.now()

# readBook = openpyxl.load_workbook(configs['extractFeaturesPath'])
readBook = openpyxl.load_workbook(configs['extractFeaturesPath'])
readSheet = readBook.active

readTestBook = openpyxl.load_workbook(configs['extractTestFeaturesPath'])
readTestSheet = readTestBook.active

def getFeatureVector():
    featureVectorList = []
    sentenceTag = []
    for i in range(readSheet.max_row - 1):

        featureVector = []
        sentenceTag.append(readSheet['E' + str(i + 2)].value)
        # 先添加段落特征
        featureVector.append(float(readSheet['A' + str(i + 2)].value.strip().split('-')[1]))
        for j in range(readSheet.max_column - 5):
            if isinstance(readSheet.cell(row = i + 2, column = j + 6).value, str):
                featureVector.append(float(readSheet.cell(row=i + 2, column=j + 6).value))
            else:
                featureVector.append(readSheet.cell(row = i + 2, column = j + 6).value)
        featureVectorList.append(normalizeFeatureByMaxMin(featureVector))
    return featureVectorList, sentenceTag

# (0,1)标准化 0.389671361502
# ('4-5', 21), ('1-5', 15), ('12-5', 15), ('10-5', 12), ('11-5', 11), ('9-5', 9),
# ('5-6', 7), ('6-5', 7), ('7-5', 5), ('2-5', 5), ('6-1', 5), ('5-1', 4), ('2-6', 3),
# ('7-6', 2)
def getTestFeatureVector():
    featureVectorList = []
    sentenceTag = []
    for i in range(readTestSheet.max_row - 1):

        featureVector = []
        sentenceTag.append(readTestSheet['E' + str(i + 2)].value)
        # 先添加段落特征
        featureVector.append(float(readTestSheet['A' + str(i + 2)].value.strip().split('-')[1]))
        for j in range(readTestSheet.max_column - 5):
            if isinstance(readTestSheet.cell(row = i + 2, column = j + 6).value, str):
                featureVector.append(float(readTestSheet.cell(row=i + 2, column=j + 6).value))
            else:
                featureVector.append(readTestSheet.cell(row = i + 2, column = j + 6).value)
        featureVectorList.append(normalizeFeatureByMaxMin(featureVector))
    return featureVectorList, sentenceTag

# 归一化处理，使用sigmoid函数
# 2017.11.10 准确率 0.205202312139
# sigmoid函数做归一主要是把两边的一些噪声数据拉回来，不要让噪声数据影响模型效果，
# 而我们是自己提取的特征，已经经过了预处理，没有很多噪声数据
# 这就是在这种情况下使用sigmoid函数准确率低的原因
# ('6-5', 195), ('1-5', 108), ('4-5', 95), ('10-5', 74), ('12-5', 68), ('7-5', 66),
# ('13-5', 60), ('9-5', 51), ('2-5', 45), ('11-5', 44), ('3-5', 17), ('8-5', 2)
def normalizeFeatureBySigmoid(featureVector):
    normalizeFeatureVector = []
    for feature in featureVector:
        normalizeFeatureVector.append(1.0 / (1 + math.exp(-float(feature))))
    return normalizeFeatureVector

# 归一化处理，使用(0,1)标准化
# 2017.11.10 准确率 0.401734104046->0.413294797688
# ('1-5', 90), ('4-5', 88), ('10-5', 67), ('12-5', 65), ('7-5', 52), ('9-5', 49),
#  ('11-5', 41), ('2-5', 40), ('13-6', 34), ('13-5', 25), ('7-6', 11), ('3-5', 10),
#  ('4-6', 6), ('10-6', 6), ('1-6', 6), ('3-6', 5), ('5-6', 3), ('12-6', 3), ('7-1', 3),
#  ('2-6', 3), ('9-6', 2), ('3-1', 2), ('8-5', 2), ('11-6', 2), ('2-1', 2), ('13-1', 1),
#  ('4-1', 1), ('10-1', 1), ('11-1', 1)
def normalizeFeatureByMaxMin(featureVector):
    normalizeFeatureVector = []
    maxNum = max(featureVector)
    minNum = min(featureVector)
    for feature in featureVector:
        normalizeFeatureVector.append((feature - minNum) / (maxNum - minNum))
    return normalizeFeatureVector

# 训练并测试训练集的准确度
def doTrain(X, Y):
    clf = svm.SVC()
    clf.fit(X, Y)

    Y_pred = clf.predict(X)
    print(accuracy_score(Y, Y_pred))
    print(Y)
    print(Y_pred)
    print('Y len is ', len(Y),'Y_pred len is ', len(Y_pred))

    compare = {}
    j = 0
    for i in range(len(Y)):
        if Y[i] == 5:
            j += 1
        YAndYpred = str(Y[i]) + '-' + str(Y_pred[i])
        # print(YAndYpred)
        if(Y[i] != Y_pred[i]):
            if YAndYpred not in compare:
                compare[YAndYpred] = 1
            else:
                compare[YAndYpred] = compare[YAndYpred] + 1

    print(sorted(compare.items(), key=lambda d: -d[1]))

# 测试测试集集的准确度
def doTrainByTestSet(train_X, train_Y, test_X, test_Y):
    clf = svm.SVC()
    clf.fit(train_X, train_Y)

    Y_pred = clf.predict(test_X)
    print(accuracy_score(test_Y, Y_pred))
    print(test_Y)
    print(Y_pred)
    print('test_Y len is ', len(test_Y),'Y_pred len is ', len(Y_pred))

    compare = {}
    j = 0
    for i in range(len(test_Y)):
        YAndYpred = str(test_Y[i]) + '-' + str(Y_pred[i])
        # print(YAndYpred)
        if(test_Y[i] != Y_pred[i]):
            if YAndYpred not in compare:
                compare[YAndYpred] = 1
            else:
                compare[YAndYpred] = compare[YAndYpred] + 1

    print(sorted(compare.items(), key=lambda d: -d[1]))

params = getFeatureVector()
testParams = getTestFeatureVector()
# doTrain(params[0], params[1])
doTrainByTestSet(params[0], params[1], testParams[0], testParams[1])

# 计算程序运行总时间(秒)
elapsed = (datetime.now() - start).seconds
print('Time used : ', elapsed)